import tkinter
import cv2
import cvzone
import math
import openpyxl
from ultralytics import YOLO
from sort import *
from datetime import *

# Raiz del botón guardar
raiz = tkinter.Tk()
raiz.geometry("640x580")
raiz.title("Guardar")

# Función para guardar el contador
def guardar_contadores(contadores):
    # Abrir el archivo de excel existente o crear uno
    try:
        libro = openpyxl.load_workbook("contadores.xlsx")
        hoja = libro.active
    except FileNotFoundError:
        libro = openpyxl.workbook()
        hoja = libro.active
        hoja.title = "Contadores"

    # Escribir los datos en la siguiente fila vacía
    fila_vacia = hoja.max_row + 1
    for i, contador in enumerate(contadores):
        hoja.cell(row=fila_vacia, column=i+1, value=contador)

    libro.save("cortitelas.xlsx")


# inicializar la cámara web
cap = cv2.VideoCapture(0)

# Medidas de la ventana de la cámara
cap.set(3, 1280)
cap.set(4, 720)

# Capturar la fecha actual
today = date.today()

# Cargar el modelo de yolo que identificará los objetos
model = YOLO("../Yolo-Weights/yolov8n.pt")

# Lista con nombres de distintos objetos que reconoce yolo
classNames = ["person", "bicycle", "car", "motorbike", "aeroplane", "bus", "train", "truck", "boat",
              "traffic light", "fire hydrant", "stop sign", "parking meter", "bench", "bird", "cat",
              "dog", "horse", "sheep", "cow", "elephant", "bear", "zebra", "giraffe", "backpack", "umbrella",
              "handbag", "tie", "suitcase", "frisbee", "skis", "snowboard", "sports ball", "kite", "baseball bat",
              "baseball glove", "skateboard", "surfboard", "tennis racket", "bottle", "wine glass", "cup",
              "fork", "knife", "spoon", "bowl", "banana", "apple", "sandwich", "orange", "broccoli",
              "carrot", "hot dog", "pizza", "donut", "cake", "chair", "sofa", "pottedplant", "bed",
              "diningtable", "toilet", "tvmonitor", "laptop", "mouse", "remote", "keyboard", "cell phone",
              "microwave", "oven", "toaster", "sink", "refrigerator", "book", "clock", "vase", "scissors",
              "teddy bear", "hair drier", "toothbrush"]


# inicializando las variables de conteo
conteo = []
salidas = []

# limites Horizontales
# limitsUp = [10, 0, 10, 720]                   # Entrada
# limitsDown = [1270, 0, 1270, 720]             # salida

# Coordenadas límites verticales para poder contar a la persona
limitsUp = [0, 500, 1280, 500]               # Entrada
limitsDown = [0, 700, 1280, 700]             # salida

# variable de seguimiento de objetos
trackers = Sort(max_age=20, min_hits=3, iou_threshold=0.3)

# Mientras reciba valores
while True:

    # Capturar los fotogramas de la cámara web
    ret, frame = cap.read()

    # imagen de conteo y coordenadas


    # captura el modelo del frame que capturó la cámara web y la lee en tiempo real
    results = model(frame, stream=True)

    detections = np.empty((0, 5))

    for r in results:

        # Dibujar recuadro del objeto
        boxes = r.boxes

        for box in boxes:

            # parametros del recuadro
            x1, y1, x2, y2 = box.xyxy[0]
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
            w, h = x2 - x1, y2 - y1

            # la variable cls captura los distintos objetos que se encuentran alrededor
            conf = math.ceil((box.conf[0]*100)) / 100
            cls = int(box.cls[0])
            print(cls)

            # Si el objeto es una persona, dibuja un recuadro alrededor
            if classNames[cls] == "person" and conf > 0.3:

                # cornerRect dibuja el cuadro con los parámetros anteriores
                cvzone.cornerRect(frame, (x1, y1, w, h))

                # Coloca el nombre del objetos
                cvzone.putTextRect(frame, f'{classNames[cls]} {conf}', (max(0, x1), max(35, y1)), scale=1, thickness=1)

                currentArray = np.array([x1, y1, x2, y2, conf])
                detections = np.vstack((detections, currentArray))

    resultsTracker = trackers.update(detections)

    # Dibuja una linea límite (donde está ubicada la linea del contador) (Prueba)
    # cv2.line(frame, (limitsUp[0], limitsUp[1]), (limitsUp[2], limitsUp[3]), (0, 0, 255), 5)
    # cv2.line(frame, (limitsDown[0], limitsDown[1]), (limitsDown[2], limitsDown[3]), (0, 0, 255), 5)

    for results in resultsTracker:
        # Parametros para el contador
        x1, y1, x2, y2, id = results
        x1, y1, x2, y2 = int(x1), int(x1), int(x2), int(y2)
        print(id)

        w, h = x2 - x1, y2 - y1

        cx, cy = x1 + w // 2, y1 + h // 2

        # Dibuja un punto central en el recuadro de la persona
        cv2.circle(frame, (cx, cy), 5, (255, 0, 255), cv2.FILLED)

        # si la persona pasa la linea límite se suma al contador de personas que entran
        if limitsUp[0] < cx < limitsUp[2] and limitsUp[1] - 15 < cy < limitsUp[1] + 15:
            if conteo.count(id) == 0:
                conteo.append(id)
                cv2.line(frame, (limitsUp[3], limitsUp[2]), (limitsUp[1], limitsUp[0]), (0, 255, 0), 5)
        # si la persona pasa la linea límite se suma al contador de personas que salen
        if limitsDown[0] < cx < limitsDown[2] and limitsDown[1] - 15 < cy < limitsDown[1] + 15:
            if salidas.count(id) == 0:
                salidas.append(id)
                cv2.line(frame, (limitsDown[0], limitsDown[1]), (limitsDown[2], limitsDown[3]), (0, 255, 0), 5)

    # Imprime el conteo de personas
    cv2.putText(frame, str(len(conteo)), (895, 115), cv2.FONT_HERSHEY_PLAIN, 5, (139, 195, 75), 7)
    cv2.putText(frame, str(len(salidas)), (1150, 115), cv2.FONT_HERSHEY_PLAIN, 5, (50, 50, 230), 7)

    # Imprime la fecha actual

    cv2.putText(frame, str(today), (15, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (139, 195, 20), thickness=3)

    print(conteo)
    print(salidas)


    cv2.imshow("Image", frame)
    cv2.waitKey(1)

